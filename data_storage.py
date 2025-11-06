"""
ماژول ذخیره اطلاعات مشتریان در Excel و Google Sheets
"""

import os
import json
from datetime import datetime
from typing import Dict, Optional

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("⚠️  pandas نصب نشده است. از روش جایگزین استفاده می‌شود.")

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl نصب نشده است.")

class CustomerDataStorage:
    """کلاس برای ذخیره اطلاعات مشتریان"""
    
    def __init__(self, excel_file: str = "customers_data.xlsx", google_sheet_id: Optional[str] = None):
        """
        Initialize data storage
        
        Args:
            excel_file: مسیر فایل Excel
            google_sheet_id: ID گوگل شیت (اختیاری)
        """
        self.excel_file = excel_file
        self.google_sheet_id = google_sheet_id or os.getenv('GOOGLE_SHEET_ID')
        self.ensure_excel_file()
    
    def ensure_excel_file(self):
        """ایجاد فایل Excel در صورت عدم وجود"""
        if not os.path.exists(self.excel_file):
            if PANDAS_AVAILABLE:
                df = pd.DataFrame(columns=[
                    'شماره مشتری',
                    'تاریخ و زمان',
                    'نام و نام خانوادگی',
                    'شماره تماس',
                    'ایمیل',
                    'آدرس',
                    'محصول مورد نظر',
                    'تعداد',
                    'قیمت',
                    'وضعیت',
                    'یادداشت',
                    'Session ID'
                ])
                df.to_excel(self.excel_file, index=False, engine='openpyxl')
            else:
                # روش جایگزین بدون pandas
                wb = Workbook()
                ws = wb.active
                ws.append([
                    'شماره مشتری',
                    'تاریخ و زمان',
                    'نام و نام خانوادگی',
                    'شماره تماس',
                    'ایمیل',
                    'آدرس',
                    'محصول مورد نظر',
                    'تعداد',
                    'قیمت',
                    'وضعیت',
                    'یادداشت',
                    'Session ID'
                ])
                wb.save(self.excel_file)
            self.format_excel_file()
    
    def format_excel_file(self):
        """فرمت‌دهی فایل Excel"""
        if not OPENPYXL_AVAILABLE:
            return
            
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # فرمت هدر
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # تنظیم عرض ستون‌ها
            column_widths = {
                'A': 15,  # شماره مشتری
                'B': 20,  # تاریخ و زمان
                'C': 25,  # نام
                'D': 15,  # شماره تماس
                'E': 30,  # ایمیل
                'F': 40,  # آدرس
                'G': 30,  # محصول
                'H': 10,  # تعداد
                'I': 15,  # قیمت
                'J': 15,  # وضعیت
                'K': 40,  # یادداشت
                'L': 30,  # Session ID
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            wb.save(self.excel_file)
        except Exception as e:
            print(f"خطا در فرمت‌دهی فایل Excel: {e}")
    
    def save_customer_data(self, customer_data: Dict) -> bool:
        """
        ذخیره اطلاعات مشتری در Excel
        
        Args:
            customer_data: دیکشنری شامل اطلاعات مشتری
            
        Returns:
            True در صورت موفقیت
        """
        try:
            new_row = {
                'شماره مشتری': customer_data.get('customer_number', ''),
                'تاریخ و زمان': datetime.now().strftime('%Y/%m/%d %H:%M:%S'),
                'نام و نام خانوادگی': customer_data.get('name', ''),
                'شماره تماس': customer_data.get('phone', ''),
                'ایمیل': customer_data.get('email', ''),
                'آدرس': customer_data.get('address', ''),
                'محصول مورد نظر': customer_data.get('product', ''),
                'تعداد': customer_data.get('quantity', ''),
                'قیمت': customer_data.get('price', ''),
                'وضعیت': customer_data.get('status', 'در انتظار'),
                'یادداشت': customer_data.get('notes', ''),
                'Session ID': customer_data.get('session_id', '')
            }
            
            if PANDAS_AVAILABLE:
                # استفاده از pandas
                df = pd.read_excel(self.excel_file, engine='openpyxl')
                df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                df.to_excel(self.excel_file, index=False, engine='openpyxl')
            else:
                # روش جایگزین بدون pandas
                wb = load_workbook(self.excel_file)
                ws = wb.active
                ws.append([
                    new_row['شماره مشتری'],
                    new_row['تاریخ و زمان'],
                    new_row['نام و نام خانوادگی'],
                    new_row['شماره تماس'],
                    new_row['ایمیل'],
                    new_row['آدرس'],
                    new_row['محصول مورد نظر'],
                    new_row['تعداد'],
                    new_row['قیمت'],
                    new_row['وضعیت'],
                    new_row['یادداشت'],
                    new_row['Session ID']
                ])
                wb.save(self.excel_file)
            
            self.format_excel_file()
            
            # ذخیره در Google Sheets (اگر تنظیم شده باشد)
            if self.google_sheet_id:
                self.save_to_google_sheets(new_row)
            
            return True
            
        except Exception as e:
            print(f"خطا در ذخیره اطلاعات: {e}")
            return False
    
    def save_to_google_sheets(self, data: Dict):
        """ذخیره در Google Sheets"""
        try:
            import gspread
            from google.oauth2.service_account import Credentials
            
            # بررسی وجود credentials
            creds_file = os.getenv('GOOGLE_CREDENTIALS_FILE', 'credentials.json')
            if not os.path.exists(creds_file):
                print("⚠️  فایل credentials.json برای Google Sheets یافت نشد")
                return
            
            # اتصال به Google Sheets
            scope = ['https://spreadsheets.google.com/feeds',
                    'https://www.googleapis.com/auth/drive']
            creds = Credentials.from_service_account_file(creds_file, scopes=scope)
            client = gspread.authorize(creds)
            
            # باز کردن شیت
            sheet = client.open_by_key(self.google_sheet_id).sheet1
            
            # اضافه کردن ردیف
            row = [
                data.get('شماره مشتری', ''),
                data.get('تاریخ و زمان', ''),
                data.get('نام و نام خانوادگی', ''),
                data.get('شماره تماس', ''),
                data.get('ایمیل', ''),
                data.get('آدرس', ''),
                data.get('محصول مورد نظر', ''),
                data.get('تعداد', ''),
                data.get('قیمت', ''),
                data.get('وضعیت', ''),
                data.get('یادداشت', ''),
                data.get('Session ID', '')
            ]
            sheet.append_row(row)
            
            print("✅ اطلاعات در Google Sheets ذخیره شد")
            
        except Exception as e:
            print(f"⚠️  خطا در ذخیره در Google Sheets: {e}")
    
    def get_all_customers(self):
        """دریافت تمام اطلاعات مشتریان"""
        try:
            if PANDAS_AVAILABLE:
                return pd.read_excel(self.excel_file, engine='openpyxl')
            else:
                # روش جایگزین
                wb = load_workbook(self.excel_file)
                ws = wb.active
                data = []
                for row in ws.iter_rows(values_only=True):
                    data.append(row)
                return data
        except Exception as e:
            print(f"خطا در خواندن فایل: {e}")
            if PANDAS_AVAILABLE:
                return pd.DataFrame()
            return []

